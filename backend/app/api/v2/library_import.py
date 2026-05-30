"""
CSV / XLSX → user_formulas bulk import — Phase 9.3 of BUILD_ROADMAP.md.

Two endpoints:

  POST /api/v2/library/import/preview
       multipart/form-data:
           file        the CSV or XLSX upload (≤ 5 MB)
       form:
           user_id     the owning auth.users.uuid (the Worker adds it)
       returns 200 with
           { ok, total_rows, valid_rows, rows: [...], errors: [...] }

  POST /api/v2/library/import/commit
       JSON:
           { user_id, rows: [<validated rows from preview>] }
       returns 200 with
           { ok, inserted, ids: [...] }

The Worker is the public auth gate (it verifies the JWT and binds the
user_id before forwarding); both endpoints additionally check the
internal shared-secret header so they can't be hit directly from the
public internet.

File format
-----------
One row per formula. Components inside one cell as ingredients-list
using a pipe (`|`) for fields and a semicolon (`;`) for items:

    Water | 75 | 7732-18-5 | solvent ; SLS | 12 | 151-21-3 | surfactant

Required columns:
    name
    ingredients   (the pipe/semicolon string above)
Optional columns:
    name_en, category, sub_category, form_type, description,
    notes, trust_score, project, tags

Tags can be comma-separated in one cell:  "wip, herbal, R&D"

Anything beyond these column names is ignored — chemists can keep their
own metadata in extra columns without breaking the parser.
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from pydantic import BaseModel

from openpyxl import load_workbook  # already in backend/requirements.txt

router = APIRouter(prefix="/library", tags=["library-import"])
log = logging.getLogger("library_import")

MAX_BYTES = 5 * 1024 * 1024   # 5 MB upload cap — keeps Workers happy
MAX_ROWS  = 2000              # one upload, one workspace; not a data lake

CANONICAL_COLUMNS = {
    "name", "name_en", "category", "sub_category", "form_type",
    "description", "notes", "trust_score", "project", "tags",
    "ingredients",
}
REQUIRED_COLUMNS = {"name", "ingredients"}


# ── helpers ──────────────────────────────────────────────────────────


def _norm_header(h: str) -> str:
    """Make the column header lookup forgiving — trim + lowercase +
    collapse underscores/spaces. So "Sub Category", "sub_category",
    " sub-category " all match `sub_category`."""
    return re.sub(r"[\s_-]+", "_", (h or "").strip().lower())


def _require_internal(request: Request) -> None:
    expected = os.getenv("BACKEND_INTERNAL_SECRET", "missing-secret")
    if request.headers.get("x-formula-internal") != expected:
        raise HTTPException(status_code=403, detail="forbidden")


def _parse_ingredients(cell: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Pipe/semicolon ingredient list → list of component dicts.

    Returns (components, errors_for_this_cell).
    """
    if not cell or not str(cell).strip():
        return [], ["ingredients cell is empty"]

    items = [chunk.strip() for chunk in str(cell).split(";") if chunk.strip()]
    if not items:
        return [], ["no ingredient rows parsed"]

    components: list[dict[str, Any]] = []
    errors: list[str] = []
    for i, chunk in enumerate(items, 1):
        parts = [p.strip() for p in chunk.split("|")]
        if len(parts) < 2:
            errors.append(f"ingredient {i}: expected `name | pct | cas? | function?`, got {chunk!r}")
            continue
        name = parts[0]
        try:
            pct = float(parts[1])
        except (TypeError, ValueError):
            errors.append(f"ingredient {i} ({name!r}): percentage {parts[1]!r} is not a number")
            continue
        cas = parts[2] if len(parts) >= 3 and parts[2] else None
        func = parts[3] if len(parts) >= 4 and parts[3] else None
        components.append({
            "name_en": name,
            "name": name,
            "percentage": pct,
            "cas_number": cas,
            "function": func,
        })
    return components, errors


def _parse_tags(cell: Any) -> list[str]:
    if not cell:
        return []
    return [t.strip() for t in str(cell).split(",") if t.strip()]


def _parse_trust_score(cell: Any) -> int | None:
    if cell is None or cell == "":
        return None
    try:
        v = int(float(str(cell).strip()))
        return max(0, min(100, v))
    except (TypeError, ValueError):
        return None


def _row_to_formula(row: dict[str, Any], row_n: int) -> tuple[dict[str, Any] | None, list[str]]:
    """Pick the columns we know about, validate, return (payload, errors).

    `payload` is None when the row is so broken we shouldn't even offer
    it to the commit step. Otherwise the caller can decide whether to
    surface the errors as warnings or strict failures.
    """
    errs: list[str] = []

    name = str(row.get("name") or "").strip()
    if not name:
        errs.append("missing required `name`")

    ingredients_cell = row.get("ingredients") or ""
    components, ing_errs = _parse_ingredients(ingredients_cell)
    errs.extend(ing_errs)
    if not components:
        # No usable components — refuse to insert this row.
        return None, errs or ["no usable ingredients"]

    payload = {
        "name": name[:200] if name else f"Imported row {row_n}",
        "name_en": (str(row.get("name_en") or name)).strip()[:200] or None,
        "category": (str(row.get("category") or "").strip() or None),
        "sub_category": (str(row.get("sub_category") or "").strip() or None),
        "form_type": (str(row.get("form_type") or "").strip() or None),
        "description": (str(row.get("description") or "").strip() or None),
        "notes": (str(row.get("notes") or "").strip() or None),
        "components": components,
        "trust_score": _parse_trust_score(row.get("trust_score")) or 80,
        "project": (str(row.get("project") or "").strip() or None),
        "tags": _parse_tags(row.get("tags")),
    }
    return payload, errs


# ── file readers ────────────────────────────────────────────────────


def _read_csv(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")  # strip BOM if Excel added one
    reader = csv.DictReader(io.StringIO(text))
    out: list[dict[str, Any]] = []
    for row in reader:
        normalised = {_norm_header(k): v for k, v in row.items() if k}
        out.append(normalised)
    return out


def _read_xlsx(data: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [_norm_header(str(h) if h is not None else "") for h in next(rows_iter)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for r in rows_iter:
        row = {h: v for h, v in zip(header, r) if h}
        # Skip wholly-empty rows so the parser doesn't trip on stray
        # bottom rows users sometimes leave behind in Excel.
        if not any(str(v or "").strip() for v in row.values()):
            continue
        out.append(row)
    return out


def _parse_upload(content: bytes, filename: str | None) -> list[dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return _read_csv(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _read_xlsx(content)
    # Try CSV first, then XLSX, so older Excel/Numbers exports still work.
    try:
        return _read_csv(content)
    except Exception:
        return _read_xlsx(content)


# ── /preview ───────────────────────────────────────────────────────


class PreviewResponse(BaseModel):
    ok: bool
    total_rows: int
    valid_rows: int
    rows: list[dict[str, Any]]          # validated payloads, ready for /commit
    errors: list[dict[str, Any]]        # [{row_n, errors: [str]}]
    sample_columns: list[str]           # normalised header list, for the UI
    notes: list[str] = []


@router.post("/import/preview", response_model=PreviewResponse)
async def import_preview(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Form(...),
):
    _require_internal(request)

    if not user_id or len(user_id) < 8:
        raise HTTPException(status_code=400, detail="missing user_id")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")
    if len(content) > MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"file too large ({len(content):,} bytes > {MAX_BYTES:,})",
        )

    try:
        raw_rows = _parse_upload(content, file.filename)
    except Exception as exc:  # noqa: BLE001
        log.exception("upload parse failed")
        raise HTTPException(status_code=422, detail=f"could not parse file: {exc}") from exc

    if not raw_rows:
        raise HTTPException(status_code=422, detail="file has no data rows")
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=413,
            detail=f"too many rows ({len(raw_rows)} > {MAX_ROWS}); split the upload",
        )

    cols = sorted(raw_rows[0].keys()) if raw_rows else []
    missing_required = REQUIRED_COLUMNS - set(cols)
    if missing_required:
        raise HTTPException(
            status_code=422,
            detail=f"missing required columns: {', '.join(sorted(missing_required))}",
        )

    valid_payloads: list[dict[str, Any]] = []
    row_errors: list[dict[str, Any]] = []
    for i, row in enumerate(raw_rows, start=2):  # row 1 = header
        payload, errs = _row_to_formula(row, i)
        if payload is not None:
            valid_payloads.append(payload)
            if errs:
                row_errors.append({"row_n": i, "errors": errs, "severity": "warning"})
        else:
            row_errors.append({"row_n": i, "errors": errs, "severity": "skipped"})

    notes: list[str] = []
    ignored_cols = sorted(set(cols) - CANONICAL_COLUMNS)
    if ignored_cols:
        notes.append(f"ignored columns (kept in your spreadsheet, not stored): {', '.join(ignored_cols)}")

    return PreviewResponse(
        ok=True,
        total_rows=len(raw_rows),
        valid_rows=len(valid_payloads),
        rows=valid_payloads,
        errors=row_errors,
        sample_columns=cols,
        notes=notes,
    )


# ── /commit ────────────────────────────────────────────────────────


class CommitRequest(BaseModel):
    user_id: str
    rows: list[dict[str, Any]]


class CommitResponse(BaseModel):
    ok: bool
    inserted: int
    ids: list[str]


@router.post("/import/commit", response_model=CommitResponse)
async def import_commit(request: Request, body: CommitRequest):
    _require_internal(request)

    if not body.user_id or len(body.user_id) < 8:
        raise HTTPException(status_code=400, detail="missing user_id")
    if not body.rows:
        raise HTTPException(status_code=400, detail="empty rows")
    if len(body.rows) > MAX_ROWS:
        raise HTTPException(status_code=413, detail="too many rows")

    supabase = request.app.state.supabase

    # Stamp user_id on every row before insert. RLS would catch a missing
    # user_id with service_role, but we set it explicitly so we don't
    # rely on the policy as our only check.
    payloads = []
    for r in body.rows:
        p = dict(r)
        p["user_id"] = body.user_id
        # Strip any tags=[] so the column default kicks in cleanly.
        if not p.get("tags"):
            p.pop("tags", None)
        payloads.append(p)

    res = supabase.table("user_formulas").insert(payloads).execute()
    ids = [row.get("id") for row in (res.data or []) if isinstance(row, dict) and row.get("id")]
    return CommitResponse(ok=True, inserted=len(ids), ids=ids)
